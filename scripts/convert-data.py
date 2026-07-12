from __future__ import annotations

import json
import math
import os
import shutil
import csv
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
FORECAST_PROJECT_DIR = Path(
    os.getenv("FX_FORECAST_PROJECT_DIR", str(ROOT.parent / "02 Prophet退火算法 20260201"))
).resolve()
DEFAULT_UPLOAD_DIR = (
    FORECAST_PROJECT_DIR / "upload file"
    if (FORECAST_PROJECT_DIR / "upload file").exists()
    else ROOT / "upload file"
)
UPLOAD_DIR = Path(os.getenv("FX_FORECAST_UPLOAD_DIR", str(DEFAULT_UPLOAD_DIR))).resolve()
TERRAIN_SNAPSHOT_PATH = Path(
    os.getenv(
        "FX_TERRAIN_SNAPSHOT_PATH",
        str(FORECAST_PROJECT_DIR / "prophet output" / "terrain" / "latest_terrain_snapshot.xlsx"),
    )
).resolve()
TRADE_SIGNAL_PATH = Path(
    os.getenv(
        "FX_TRADE_SIGNAL_PATH",
        str(FORECAST_PROJECT_DIR / "prophet output" / "trade signals" / "latest_trade_signals.xlsx"),
    )
).resolve()
TERRAIN_FEATURE_DIR = Path(
    os.getenv(
        "FX_TERRAIN_FEATURE_DIR",
        str(FORECAST_PROJECT_DIR / "prophet output" / "terrain"),
    )
).resolve()
DATA_DIR = ROOT / "public" / "data"
FILES_DIR = DATA_DIR / "files"
TERRAIN_DATA_DIR = DATA_DIR / "terrain"

GROUPS = {
    "人民币相关": ["USDCNH", "EURCNH", "GBPCNH", "AUDCNH"],
    "美元直盘": ["EURUSD", "GBPUSD", "AUDUSD", "USDJPY", "USDCHF"],
    "贵金属": ["XAUUSD"],
}

NAMES = {
    "USDCNH": "美元 / 离岸人民币",
    "EURCNH": "欧元 / 离岸人民币",
    "GBPCNH": "英镑 / 离岸人民币",
    "AUDCNH": "澳元 / 离岸人民币",
    "EURUSD": "欧元 / 美元",
    "GBPUSD": "英镑 / 美元",
    "AUDUSD": "澳元 / 美元",
    "USDJPY": "美元 / 日元",
    "USDCHF": "美元 / 瑞郎",
    "XAUUSD": "黄金 / 美元",
}


def read_two_column_xlsx(path: Path) -> list[tuple[str, float]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows: list[tuple[str, float]] = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        raw_date, raw_value = row[:2]
        if raw_date is None or raw_value is None:
            continue
        if isinstance(raw_date, datetime):
            date = raw_date.date().isoformat()
        else:
            date = str(raw_date)[:10]
        rows.append((date, float(raw_value)))

    return rows


def pct_band(symbol: str, value: float) -> float:
    if symbol == "XAUUSD":
        return max(value * 0.035, 25)
    if symbol == "USDJPY":
        return max(value * 0.018, 1.2)
    if symbol.endswith("CNH"):
        return max(value * 0.015, 0.045)
    return max(value * 0.018, 0.006)


def fmt(value: float) -> float:
    if abs(value) >= 100:
        return round(value, 2)
    if abs(value) >= 10:
        return round(value, 4)
    return round(value, 5)


def json_value(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return None
        return value
    return str(value)


def read_records_by_symbol(path: Path) -> dict[str, dict]:
    if not path.exists():
        print(f"Optional strategy data not found: {path}")
        return {}
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value) if value is not None else "" for value in next(rows, [])]
    records: dict[str, dict] = {}
    for values in rows:
        record = {header: json_value(value) for header, value in zip(headers, values) if header}
        symbol = str(record.get("Pair") or "").upper()
        if symbol:
            records[symbol] = record
    return records


def number_or_none(value):
    return value if isinstance(value, (int, float)) and math.isfinite(float(value)) else None


def csv_number(value):
    try:
        number = float(value)
        return number if math.isfinite(number) else None
    except (TypeError, ValueError):
        return None


def read_terrain_series(symbol: str, limit: int = 1040) -> list[dict]:
    path = TERRAIN_FEATURE_DIR / f"{symbol}_terrain_features.csv"
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))[-limit:]
    fast_periods = [3, 5, 8, 10, 12, 15]
    slow_periods = [30, 35, 40, 45, 50, 60]
    series = []
    for row in rows:
        close = csv_number(row.get("Close"))
        if not row.get("Date") or close is None:
            continue
        fast = [csv_number(row.get(f"EMA_{period}")) for period in fast_periods]
        slow = [csv_number(row.get(f"EMA_{period}")) for period in slow_periods]
        d_layers = [csv_number(row.get(f"D{index}")) for index in range(1, 7)]
        area_layers = [csv_number(row.get(f"Area_{index}")) for index in range(1, 7)]
        if any(value is None for value in fast + slow + d_layers + area_layers):
            continue
        trend = csv_number(row.get("Trend"))
        coherence = csv_number(row.get("coh"))
        regime_age = csv_number(row.get("RegimeAge"))
        series.append({
            "date": row["Date"][:10],
            "close": fmt(close),
            "fast": [fmt(value) for value in fast],
            "slow": [fmt(value) for value in slow],
            "trend": int(trend) if trend is not None else 0,
            "gate": round(csv_number(row.get("TerrainGate")) or 0.0, 6),
            "score": round(csv_number(row.get("TrendScore")) or 0.0, 6),
            "energyAtr": round(csv_number(row.get("EnergyATR")) or 0.0, 6),
            "atr": fmt(csv_number(row.get("ATR14")) or 0.0),
            "coherence": int(coherence) if coherence is not None else 0,
            "regimeAge": int(regime_age) if regime_age is not None else 0,
            "d": [fmt(value) for value in d_layers],
            "area": [round(value, 6) for value in area_layers],
        })
    return series


def build_terrain_payload(record: dict | None) -> dict | None:
    if not record:
        return None
    return {
        "date": record.get("TerrainDate"),
        "state": record.get("TerrainState"),
        "trend": number_or_none(record.get("TerrainTrend")),
        "gate": number_or_none(record.get("TerrainGate")),
        "trendScore": number_or_none(record.get("TerrainTrendScore")),
        "threshold": number_or_none(record.get("TerrainThreshold")),
        "band": number_or_none(record.get("TerrainBand")),
        "coherence": number_or_none(record.get("TerrainCoherence")),
        "coherenceRatio": number_or_none(record.get("TerrainCoherenceRatio")),
        "regimeId": number_or_none(record.get("TerrainRegimeID")),
        "regimeAge": number_or_none(record.get("TerrainRegimeAge")),
        "energy": number_or_none(record.get("TerrainEnergy")),
        "energyAtr": number_or_none(record.get("TerrainEnergyATR")),
        "bundleOverlap": number_or_none(record.get("TerrainBundleOverlap")),
        "bundleSep": number_or_none(record.get("TerrainBundleSep")),
        "filterEnabled": bool(record.get("FilterEnabled")),
    }


def build_trade_payload(record: dict | None) -> dict | None:
    if not record:
        return None
    number_fields = {
        "signalHorizon": "SignalHorizon",
        "terrainGate": "TerrainGate",
        "terrainTrendScore": "TerrainTrendScore",
        "terrainCoherence": "TerrainCoherence",
        "terrainRegimeAge": "TerrainRegimeAge",
        "sizeMultiplier": "TerrainSizeMultiplier",
        "entryPrice": "EntryPrice",
        "targetPrice": "TargetPrice",
        "takeProfit": "TakeProfit",
        "stopLoss": "StopLoss",
        "returnPct": "ReturnPct",
        "riskReward": "RiskReward",
        "candidateNotional": "CandidateNotional",
        "positionNotional": "PositionNotional",
        "marginPct": "MarginPct",
        "maxLossPct": "MaxLossPct",
    }
    payload = {key: number_or_none(record.get(source)) for key, source in number_fields.items()}
    payload.update({
        "runDate": record.get("RunDate"),
        "engine": record.get("Engine"),
        "decisionPolicy": record.get("DecisionPolicy"),
        "decisionConfidence": record.get("DecisionConfidence"),
        "rawDirection": record.get("RawDirection"),
        "direction": record.get("Direction"),
        "terrainState": record.get("TerrainState"),
        "terrainDate": record.get("TerrainDate"),
        "alignment": record.get("TerrainAlignment"),
        "action": record.get("TerrainAction"),
    })
    return payload


def build_symbol(symbol: str, terrain_records: dict[str, dict], trade_records: dict[str, dict]) -> dict:
    history_path = UPLOAD_DIR / f"{symbol}_diff_0th_diff.xlsx"
    forecast_path = UPLOAD_DIR / f"{symbol}_forecast.xlsx"

    history = read_two_column_xlsx(history_path)
    forecast = read_two_column_xlsx(forecast_path)
    actual_by_date = dict(history)
    forecast_by_date = dict(forecast)
    dates = sorted(set(actual_by_date) | set(forecast_by_date))

    latest_actual_date, latest_actual_value = history[-1]
    future_forecasts = [(date, value) for date, value in forecast if date > latest_actual_date]
    next_forecast_date, next_forecast_value = future_forecasts[0] if future_forecasts else forecast[-1]
    deviation = next_forecast_value - latest_actual_value
    direction = "上行" if deviation > 0 else "下行" if deviation < 0 else "持平"

    series = []
    for date in dates:
        actual = actual_by_date.get(date)
        predicted = forecast_by_date.get(date)
        base = predicted if predicted is not None else actual
        low = high = None
        if base is not None and predicted is not None:
            band = pct_band(symbol, base)
            low = base - band
            high = base + band
        series.append(
            {
                "date": date,
                "actual": fmt(actual) if actual is not None else None,
                "forecast": fmt(predicted) if predicted is not None else None,
                "lower": fmt(low) if low is not None else None,
                "upper": fmt(high) if high is not None else None,
                "future": date > latest_actual_date,
            }
        )

    trade_signal = build_trade_payload(trade_records.get(symbol))
    terrain = build_terrain_payload(terrain_records.get(symbol))
    return {
        "symbol": symbol,
        "name": NAMES.get(symbol, symbol),
        "group": next(group for group, symbols in GROUPS.items() if symbol in symbols),
        "frequency": "周频",
        "modelVersion": "Prophet + Terrain v1" if trade_signal else "Prophet annealing v1",
        "generatedAt": datetime.now(timezone.utc).date().isoformat(),
        "sourceUpdatedAt": max(history_path.stat().st_mtime, forecast_path.stat().st_mtime),
        "latestActual": {"date": latest_actual_date, "value": fmt(latest_actual_value)},
        "nextForecast": {"date": next_forecast_date, "value": fmt(next_forecast_value)},
        "direction": direction,
        "deviation": fmt(deviation),
        "terrain": terrain,
        "terrainSeriesUrl": f"/data/terrain/{symbol}.json",
        "tradeSignal": trade_signal,
        "files": {
            "forecast": f"/data/files/{forecast_path.name}",
            "history": f"/data/files/{history_path.name}",
            "terrain": f"/data/files/{TERRAIN_SNAPSHOT_PATH.name}" if TERRAIN_SNAPSHOT_PATH.exists() else None,
            "tradeSignals": f"/data/files/{TRADE_SIGNAL_PATH.name}" if TRADE_SIGNAL_PATH.exists() else None,
        },
        "series": series,
    }


def main() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    FILES_DIR.mkdir(parents=True, exist_ok=True)
    TERRAIN_DATA_DIR.mkdir(parents=True, exist_ok=True)

    all_symbols = [symbol for symbols in GROUPS.values() for symbol in symbols]
    manifest_symbols = []
    terrain_records = read_records_by_symbol(TERRAIN_SNAPSHOT_PATH)
    trade_records = read_records_by_symbol(TRADE_SIGNAL_PATH)

    for symbol in all_symbols:
        data = build_symbol(symbol, terrain_records, trade_records)
        terrain_series = read_terrain_series(symbol)
        (TERRAIN_DATA_DIR / f"{symbol}.json").write_text(
            json.dumps({"symbol": symbol, "series": terrain_series}, ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8",
        )
        updated = datetime.fromtimestamp(data["sourceUpdatedAt"]).isoformat(timespec="seconds")
        data["sourceUpdatedAt"] = updated
        (DATA_DIR / f"{symbol}.json").write_text(
            json.dumps(data, ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8",
        )

        for file_name in [f"{symbol}_forecast.xlsx", f"{symbol}_diff_0th_diff.xlsx"]:
            shutil.copy2(UPLOAD_DIR / file_name, FILES_DIR / file_name)

        manifest_symbols.append(
            {
                "symbol": symbol,
                "name": NAMES.get(symbol, symbol),
                "group": data["group"],
                "latestActual": data["latestActual"],
                "nextForecast": data["nextForecast"],
                "direction": data["direction"],
                "terrain": data["terrain"],
                "tradeSignal": data["tradeSignal"],
            }
        )

    for optional_path in [TERRAIN_SNAPSHOT_PATH, TRADE_SIGNAL_PATH]:
        if optional_path.exists():
            shutil.copy2(optional_path, FILES_DIR / optional_path.name)

    manifest = {
        "title": "汇率与贵金属预测观察",
        "defaultSymbol": "USDCNH",
        "updatedAt": max(
            item.stat().st_mtime
            for item in UPLOAD_DIR.glob("*.xlsx")
            if not math.isnan(item.stat().st_mtime)
        ),
        "groups": GROUPS,
        "symbols": manifest_symbols,
    }
    manifest["updatedAt"] = datetime.fromtimestamp(manifest["updatedAt"]).isoformat(timespec="seconds")
    (DATA_DIR / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )

    print(f"Wrote {len(all_symbols)} symbols to {DATA_DIR}")


if __name__ == "__main__":
    main()
